import openpyxl as opxl
import numpy as np
import time 

def month_to_number(month):
    months = {
        "январь": ".01",
        "февраль": ".02",
        "март": ".03",
        "апрель": ".04",
        "май": ".05",
        "июнь": ".06",
        "июль": ".07",
        "август": ".08",
        "сентябрь": ".09",
        "октябрь": ".10",
        "ноябрь": ".11",
        "декабрь": ".12"
    }
    
    return months.get(month.lower(), "Некорректное название месяца")

def sort_exels(file,file_new):
    book=opxl.open(file,read_only=True)
    sheet=book.active
    data=[] # создаем массив для хранения данных за весь промежуток времени
    if(sheet.cell(row=5,column=2).value!=None):
        info=[sheet.cell(row=4,column=9).value]
        size_info=1 #количество параметров
        while(sheet.cell(row=4,column=9).value!=sheet.cell(row=4+size_info,column=9).value):
            info.append(sheet.cell(row=4+size_info,column=9).value)
            size_info+=1
    print("Start project:")
    start_time=time.time()
    start_=5
    while (sheet.cell(row=start_,column=2).value!=None):
        time_=time.time()
        month_y=sheet.cell(row=start_,column=2).value
        size=len(month_y)
        month=month_to_number(month_y[0:size-6])
        year=month_y[size-4:]
        date_m=[] # создаем массив для хранения данных для определенного месяца
        i=10
        while(sheet.cell(row=start_-1,column=i).value!=None):
            if(sheet.cell(row=start_-1,column=i).value==None):
                i=i+1
                continue
            date_d=[] # создаем массив для хранения данных для определенного дня
            date_d.append(str(sheet.cell(row=3,column=i).value)+month+"."+year)
            for k in range(0, size_info):
                date_d.append(sheet.cell(row=start_-1+k,column=i).value)
            date_m.append(date_d)
            i=i+1
        data.append(date_m)
        print("Current line:= "+str(start_)+", execution speed: "+str(time.time()-time_))
        start_+=size_info

    data.reverse()
    book_new=opxl.Workbook()
    sheet_new=book_new.active

    for i in range(0,size_info):
        sheet_new.cell(row=1,column=2+i).value=info[i]

    i=0
    for date_m_ in data:
        for date_d_ in date_m_:
            for j in range(0,size_info+1):
                sheet_new.cell(row=2+i,column=1+j).value=date_d_[j]
            i=i+1
    book_new.save(file_new)
    book_new.close()
    book.close()
    print("Finish project, all time:"+str(time.time()-start_time))


sort_exels("Скв. 9543 (01.02.2022-30.06.2023)_Газпромнефть 1.xlsx","Скв. 9543 (01.02.2022-30.06.2023)_Газпромнефть 1_new.xlsx")
# sort_exels("Скв. 388 (01.11.2016-30.11.2023).xlsx","1_new.xlsx")

